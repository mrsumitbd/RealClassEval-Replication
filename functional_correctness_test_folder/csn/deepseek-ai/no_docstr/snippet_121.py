
import openpyxl


class Excels:

    @classmethod
    def read(cls, file_name, tab):
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook[tab]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        return data

    @classmethod
    def read_opyxl(cls, file_name):
        workbook = openpyxl.load_workbook(file_name)
        sheets_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)
            sheets_data[sheet_name] = data
        return sheets_data

    @classmethod
    def clean(cls):
        pass
