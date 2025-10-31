
import pandas as pd
import openpyxl


class Excels:
    _data = None

    @classmethod
    def read(cls, file_name, tab):
        cls._data = pd.read_excel(file_name, sheet_name=tab)
        return cls._data

    @classmethod
    def read_opyxl(cls, file_name):
        wb = openpyxl.load_workbook(file_name)
        data = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                data[sheet] = []
                continue
            headers = rows[0]
            sheet_data = [dict(zip(headers, row)) for row in rows[1:]]
            data[sheet] = sheet_data
        cls._data = data
        return cls._data

    @classmethod
    def clean(cls):
        cls._data = None
