class Excels:
    _workbooks = {}

    @classmethod
    def read(cls, file_name, tab):
        wb = cls.read_opyxl(file_name)
        if isinstance(tab, int):
            try:
                sheet_name = wb.sheetnames[tab]
            except Exception as e:
                raise IndexError(f"Tab index out of range: {tab}") from e
        else:
            sheet_name = str(tab)
        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet not found: {sheet_name}")
        ws = wb[sheet_name]
        return [list(row) for row in ws.iter_rows(values_only=True)]

    @classmethod
    def read_opyxl(cls, file_name):
        from openpyxl import load_workbook
        key = str(file_name)
        if key not in cls._workbooks:
            wb = load_workbook(filename=file_name,
                               read_only=True, data_only=True)
            cls._workbooks[key] = wb
        return cls._workbooks[key]

    @classmethod
    def clean(cls):
        '''
        Clean the dictionary of read files
        '''
        for wb in cls._workbooks.values():
            try:
                wb.close()
            except Exception:
                pass
        cls._workbooks.clear()
