import os
from pathlib import Path
from threading import RLock

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


class Excels:
    '''
    Class to save the read Excel files and thus avoid double reading
    '''
    _pd_cache = {}        # {abs_path: {sheet_name: DataFrame}}
    _opyxl_cache = {}     # {abs_path: Workbook}
    _lock = RLock()

    @classmethod
    def _normalize_path(cls, file_name):
        return str(Path(file_name).expanduser().resolve())

    @classmethod
    def read(cls, file_name, tab):
        '''
        Read the Excel file or return the previously read one
        '''
        if pd is None:
            raise RuntimeError(
                "pandas is required for Excels.read but is not available")
        path = cls._normalize_path(file_name)
        with cls._lock:
            sheet_cache = cls._pd_cache.setdefault(path, {})
            if tab in sheet_cache:
                return sheet_cache[tab]
        df = pd.read_excel(path, sheet_name=tab)
        with cls._lock:
            # In case another thread populated it while reading
            sheet_cache = cls._pd_cache.setdefault(path, {})
            sheet_cache[tab] = df
            return df

    @classmethod
    def read_opyxl(cls, file_name):
        '''
        Read the Excel file using OpenPyXL or return the previously read one
        '''
        if load_workbook is None:
            raise RuntimeError(
                "openpyxl is required for Excels.read_opyxl but is not available")
        path = cls._normalize_path(file_name)
        with cls._lock:
            if path in cls._opyxl_cache:
                return cls._opyxl_cache[path]
        wb = load_workbook(path, data_only=True)
        with cls._lock:
            # In case another thread populated it while reading
            if path not in cls._opyxl_cache:
                cls._opyxl_cache[path] = wb
            return cls._opyxl_cache[path]

    @classmethod
    def clean(cls):
        '''
        Clean the dictionary of read files
        '''
        with cls._lock:
            cls._pd_cache.clear()
            cls._opyxl_cache.clear()
