class Excels:
    _cache = {}

    @classmethod
    def _ensure_path(cls, file_name):
        import os
        if not isinstance(file_name, str) or not file_name.strip():
            raise ValueError("file_name must be a non-empty string")
        if not os.path.exists(file_name):
            raise FileNotFoundError(f"File not found: {file_name}")

    @classmethod
    def _strip_str(cls, v):
        if isinstance(v, str):
            s = v.strip()
            return s
        return v

    @classmethod
    def _row_is_empty(cls, row):
        for v in row:
            if v is None or (isinstance(v, float) and v != v):
                continue
            if isinstance(v, str) and v.strip() == "":
                continue
            return False
        return True

    @classmethod
    def _to_records(cls, headers, rows):
        clean_headers = []
        seen = {}
        for h in headers:
            key = str(h).strip() if h is not None else ""
            if key == "":
                key = "column"
            base = key
            if key in seen:
                seen[key] += 1
                key = f"{base}_{seen[base]}"
            else:
                seen[key] = 0
            clean_headers.append(key)
        records = []
        for r in rows:
            if cls._row_is_empty(r):
                continue
            rec = {}
            for i, h in enumerate(clean_headers):
                v = r[i] if i < len(r) else None
                if v is None:
                    rec[h] = None
                else:
                    if isinstance(v, float) and v != v:
                        rec[h] = None
                    else:
                        rec[h] = cls._strip_str(v)
            records.append(rec)
        return records

    @classmethod
    def read(cls, file_name, tab):
        cls._ensure_path(file_name)
        cache_key = ("pandas", file_name, tab)
        if cache_key in cls._cache:
            return cls._cache[cache_key]
        try:
            import pandas as pd
            df = pd.read_excel(file_name, sheet_name=tab, dtype=object)
            # Normalize column names
            df.columns = [
                str(c).strip() if c is not None else "" for c in df.columns]
            # Replace NaN with None
            df = df.where(pd.notnull(df), None)
            # Strip strings
            for c in df.columns:
                df[c] = df[c].map(cls._strip_str)
            records = df.to_dict(orient="records")
            cls._cache[cache_key] = records
            return records
        except ImportError:
            # Fallback to openpyxl-only read
            wb_data = cls.read_opyxl(file_name)
            # tab can be index or name
            if isinstance(tab, int):
                try:
                    sheet_name = list(wb_data.keys())[tab]
                except IndexError:
                    raise ValueError(f"Sheet index out of range: {tab}")
            else:
                sheet_name = tab
            if sheet_name not in wb_data:
                raise ValueError(f"Sheet not found: {sheet_name}")
            records = wb_data[sheet_name]
            cls._cache[cache_key] = records
            return records

    @classmethod
    def read_opyxl(cls, file_name):
        cls._ensure_path(file_name)
        cache_key = ("openpyxl", file_name)
        if cache_key in cls._cache:
            return cls._cache[cache_key]
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError(
                "openpyxl is required to read Excel files without pandas")
        wb = load_workbook(filename=file_name, read_only=True, data_only=True)
        result = {}
        try:
            for ws in wb.worksheets:
                rows_iter = ws.iter_rows(values_only=True)
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    result[ws.title] = []
                    continue
                headers = [h if h is not None else "" for h in header_row]
                data_rows = []
                for r in rows_iter:
                    data_rows.append(list(r))
                records = cls._to_records(headers, data_rows)
                result[ws.title] = records
        finally:
            wb.close()
        cls._cache[cache_key] = result
        return result

    @classmethod
    def clean(cls):
        cls._cache.clear()
