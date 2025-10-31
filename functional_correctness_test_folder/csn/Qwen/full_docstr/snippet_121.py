
import openpyxl
from openpyxl import load_workbook


class Excels:
    '''
    Class to save the read Excel files and thus avoid double reading
    '''
    _cache = {}

    @classmethod
    def read(cls, file_name, tab):
        '''
        Read the Excel file or return the previously read one
        '''
        key = (file_name, tab)
        if key not in cls._cache:
            workbook = load_workbook(filename=file_name, data_only=True)
            cls._cache[key] = workbook[tab]
        return cls._cache[key]

    @classmethod
    def read_opyxl(cls, file_name):
        '''
        Read the Excel file using OpenPyXL or return the previously read one
        '''
        if file_name not in cls._cache:
            cls._cache[file_name] = load_workbook(
                filename=file_name, data_only=True)
        return cls._cache[file_name]

    @classmethod
    def clean(cls):
        '''
        Clean the dictionary of read files
        '''
        cls._cache.clear()
