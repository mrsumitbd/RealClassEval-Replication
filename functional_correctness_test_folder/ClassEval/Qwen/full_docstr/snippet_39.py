
import openpyxl


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        try:
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            data = [tuple(cell.value for cell in row)
                    for row in sheet.iter_rows(values_only=True)]
            return data
        except Exception as e:
            print(f"Error reading file {file_name}: {e}")
            return []

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in data:
                sheet.append(row)
            workbook.save(file_name)
            return 1
        except Exception as e:
            print(f"Error writing to file {file_name}: {e}")
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        try:
            data = self.read_excel(save_file_name)
            if not data:
                return 0, save_file_name

            processed_data = [
                tuple(cell.upper() if i == N -
                      1 else cell for i, cell in enumerate(row))
                for row in data
            ]

            success = self.write_excel(processed_data, save_file_name)
            return success, save_file_name
        except Exception as e:
            print(f"Error processing file {save_file_name}: {e}")
            return 0, save_file_name
