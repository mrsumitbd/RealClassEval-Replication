class ExcelProcessor:
    """
    This is a class for processing excel files, including readring and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        from openpyxl import load_workbook

        wb = load_workbook(filename=file_name, data_only=True)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(tuple(row))
        return data

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        >>> processor = ExcelProcessor()
        >>> new_data = [
        >>>     ('Name', 'Age', 'Country'),
        >>>     ('John', 25, 'USA'),
        >>>     ('Alice', 30, 'Canada'),
        >>>     ('Bob', 35, 'Australia'),
        >>>     ('Julia', 28, 'Germany')
        >>> ]
        >>> data = processor.write_excel(new_data, 'test_data.xlsx')
        """
        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            for row in data:
                if isinstance(row, (list, tuple)):
                    ws.append(list(row))
                else:
                    ws.append([row])
            wb.save(file_name)
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        >>> processor = ExcelProcessor()
        >>> success, output_file = processor.process_excel_data(1, 'test_data.xlsx')
        """
        from pathlib import Path

        if not isinstance(N, int) or N <= 0:
            raise ValueError(
                "N must be a positive integer (1-based column index).")

        data = self.read_excel(save_file_name)
        if not data:
            output_name = self._build_processed_filename(save_file_name)
            return 0, output_name

        col_idx = N - 1
        processed = []
        for i, row in enumerate(data):
            row_list = list(row)
            if col_idx < len(row_list) and isinstance(row_list[col_idx], str):
                if i == 0:
                    # keep header as-is
                    pass
                else:
                    row_list[col_idx] = row_list[col_idx].upper()
            processed.append(tuple(row_list))

        output_name = self._build_processed_filename(save_file_name)
        result = self.write_excel(processed, output_name)
        return result, output_name

    @staticmethod
    def _build_processed_filename(file_name):
        from pathlib import Path

        p = Path(file_name)
        stem = p.stem + "_processed"
        if p.suffix:
            return str(p.with_name(stem + p.suffix))
        else:
            return str(p.with_name(stem + ".xlsx"))
