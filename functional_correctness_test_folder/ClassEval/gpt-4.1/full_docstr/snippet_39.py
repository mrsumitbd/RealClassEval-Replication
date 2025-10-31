
import openpyxl
from openpyxl import Workbook, load_workbook
import os


class ExcelProcessor:
    """
    This is a class for processing excel files, including reading and writing excel data, as well as processing specific operations and saving as a new excel file.
    """

    def __init__(self):
        pass

    def read_excel(self, file_name):
        """
        Reading data from Excel files
        :param file_name:str, Excel file name to read
        :return:list of data, Data in Excel
        """
        data = []
        try:
            wb = load_workbook(filename=file_name)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                data.append(tuple(row))
            return data
        except Exception:
            return []

    def write_excel(self, data, file_name):
        """
        Write data to the specified Excel file
        :param data: list, Data to be written
        :param file_name: str, Excel file name to write to
        :return: 0 or 1, 1 represents successful writing, 0 represents failed writing
        """
        try:
            wb = Workbook()
            ws = wb.active
            for row in data:
                ws.append(list(row))
            wb.save(file_name)
            return 1
        except Exception:
            return 0

    def process_excel_data(self, N, save_file_name):
        """
        Change the specified column in the Excel file to uppercase
        :param N: int, The serial number of the column that want to change
        :param save_file_name: str, source file name
        :return:(int, str), The former is the return value of write_excel, while the latter is the saved file name of the processed data
        """
        data = self.read_excel(save_file_name)
        if not data or N < 1 or N > len(data[0]):
            return (0, "")

        processed_data = []
        for i, row in enumerate(data):
            row = list(row)
            if i == 0:
                # header, do not change
                processed_data.append(tuple(row))
            else:
                idx = N - 1
                if idx < len(row) and isinstance(row[idx], str):
                    row[idx] = row[idx].upper()
                processed_data.append(tuple(row))
        base, ext = os.path.splitext(save_file_name)
        output_file = f"{base}_processed{ext}"
        result = self.write_excel(processed_data, output_file)
        return (result, output_file)
